from openpyxl import load_workbook
import datetime

currentDate = datetime.datetime.now()

wb = load_workbook('./docs/schedule.xlsx', read_only=True)
sheet = wb.get_sheet_by_name('Sheet1')
col = sheet['B1'].column

def max_row(ws):
    i = 0
    for row in ws.rows:
        i+=1
        for cell in row:
            continue
    return i

def generateDTL():
    dtl = []
    for i in range(2, max_row(sheet) + 1):
        date = sheet.cell(row=i, column=2).value
        time = sheet.cell(row=i, column=3).value
        loc  = sheet.cell(row=i, column=4).value
        dtl.append((date, time, loc))
    return dtl

def getDates():
    dtl = generateDTL()
    dates = [date[0] for date in dtl]
    return dates

def getTimes():
    dtl = generateDTL()
    times = [time[1] for time in dtl]
    return times

def getLocs():
    dtl = generateDTL()
    locs = [loc[2] for loc in dtl]
    return locs

def getDay():
    return currentDate.day

def getMonth():
    return currentDate.month

def getYear():
    return currentDate.year

def sameDay(date):
    return (getDay() == date.day and
           getMonth() == date.month and
           getYear() == date.year)


def practiceToday():
    # generate all the tuples
    dtls = generateDTL()
    for dtl in dtls:
        if sameDay(dtl[0]):
            return (True, dtl[1], dtl[2])
    return (False, None, None)
